from openpyxl import load_workbook
from pathlib import Path
import re

wb=load_workbook('attached_assets/GFF_Fraud_Arena_Question_Bank_v5_1787115151178.xlsx', data_only=True)
ws=wb['Spot the Fraud']
rows=list(ws.iter_rows(min_row=5, values_only=True))
print('SPOT QUESTIONS', len(rows))
print('by level', {level: sum(r[1] == level for r in rows) for level in range(1,11)})
print('invalid option counts', [r[0] for r in rows if any(r[i] in (None, '') for i in range(12,16))])
print('invalid select', [(r[0], r[7], r[16]) for r in rows if r[7] not in (1,2)])

# Parse existing case blocks from the authored TS source enough to compare answers/nodes.
source=Path('artifacts/fraud-arena/src/data/detective.ts').read_text()
blocks=re.split(r'\n  \{\n    id: ', source)[1:]
existing={}
for b in blocks:
    identifier=b.split(',',1)[0].strip('"')
    nodes_match=re.search(r'nodes: \[(.*?)\]',b,re.S)
    answer_match=re.search(r'answer: \[(.*?)\]',b,re.S)
    if nodes_match and answer_match:
        nodes=re.findall(r'"([^"]+)"',nodes_match.group(1))
        ans=re.findall(r'"([^"]+)"',answer_match.group(1))
        existing[identifier]=(set(nodes),ans)
fd=wb['Fraud Detective']
missing=[]; changed=[]; noexisting=[]
for r in fd.iter_rows(min_row=5,values_only=True):
    cid=str(r[0]); answers=[x.strip() for x in str(r[8]).split(',')]
    if cid not in existing:
        noexisting.append(cid); continue
    nodes,old=existing[cid]
    missing_nodes=[a for a in answers if a not in nodes]
    if missing_nodes: missing.append((cid,answers,missing_nodes))
    if old != answers: changed.append((cid,old,answers))
print('FD total', fd.max_row-4, 'existing', len(existing), 'missing IDs', noexisting)
print('FD answer nodes missing from existing graphs', missing)
print('FD answer changes', len(changed))
for c in changed[:30]: print(' ',c)
