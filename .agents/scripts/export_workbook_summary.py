from openpyxl import load_workbook
from pathlib import Path
from pprint import pprint
p=Path('attached_assets/GFF_Fraud_Arena_Question_Bank_v5_1787115151178.xlsx')
wb=load_workbook(p, data_only=True)
for name in ['Spot the Fraud', 'Fraud Detective']:
    ws=wb[name]
    print('\n###',name)
    for i,row in enumerate(ws.iter_rows(values_only=True),1):
        vals=[None if v is None else str(v).replace('\n',' / ') for v in row]
        if i <= 5:
            print(i, vals)
    print('headers:', [str(c.value) if c.value is not None else '' for c in ws[ws.min_row]])
