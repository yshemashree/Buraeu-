from openpyxl import load_workbook
from pathlib import Path
import json

wb = load_workbook('attached_assets/GFF_Fraud_Arena_Question_Bank_v5_1787115151178.xlsx', data_only=True)

def js(value):
    return json.dumps(value, ensure_ascii=False)

spot = wb['Spot the Fraud']
items=[]
for r in spot.iter_rows(min_row=5, values_only=True):
    qid, level, kind, fmt, timer, points, near, select_n, marker, news, intl, stem, *tail = r
    options = [str(v) for v in tail[:4]]
    correct = [int(s.strip()) for s in str(tail[4]).split(',')]
    hook = str(tail[9] or '')
    answer_text = ' and '.join(options[i - 1] for i in correct)
    items.append({
        'id': str(qid), 'level': int(level), 'scope': str(fmt), 'kind': str(kind).lower(),
        'selectN': int(select_n), 'stem': str(stem), 'options': options, 'correct': correct,
        'why': f"Correct answer: {answer_text}.", 'hook': hook,
    })
quiz_lines = ["// Generated from GFF_Fraud_Arena_Question_Bank_v5.xlsx. Do not edit by hand.", "import type { Question } from './quiz';", "", "export const V5_QUESTIONS: Question[] = ["]
for item in items:
    quiz_lines.append('  ' + js(item) + ',')
quiz_lines += ['];', '']
Path('.agents/outputs/quiz-v5.ts').write_text('\n'.join(quiz_lines))

fd = wb['Fraud Detective']
updates={}
for r in fd.iter_rows(min_row=5, values_only=True):
    cid, pattern, sector, title, clue1, clue2, clue3, brief, answer, nodes_to_tap, points, explanation, hook = r
    updates[str(cid)] = {
        'sector': str(sector), 'title': str(title), 'clues': [str(clue1), str(clue2), str(clue3)],
        'brief': str(brief), 'answer': [x.strip() for x in str(answer).split(',')],
        'explanation': str(explanation), 'hook': str(hook),
    }
fd_lines = ["// Generated from GFF_Fraud_Arena_Question_Bank_v5.xlsx. Preserves graph topology from detective.ts.", "import type { DetectiveCase } from './detective';", "", "export const V5_DETECTIVE_UPDATES: Record<string, Pick<DetectiveCase, 'sector' | 'title' | 'clues' | 'brief' | 'answer' | 'explanation' | 'hook'>> = "]
fd_lines += [js(updates) + ';', '']
Path('.agents/outputs/detective-v5.ts').write_text('\n'.join(fd_lines))
print('quiz questions', len(items), 'chars', Path('.agents/outputs/quiz-v5.ts').stat().st_size)
print('detective updates', len(updates), 'chars', Path('.agents/outputs/detective-v5.ts').stat().st_size)
