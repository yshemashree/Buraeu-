from openpyxl import load_workbook
from pathlib import Path
p=Path('attached_assets/GFF_Fraud_Arena_Question_Bank_v5_1787115151178.xlsx')
wb=load_workbook(p, data_only=True)
print('SHEETS', wb.sheetnames)
for ws in wb.worksheets:
    print(f'\n=== {ws.title} rows={ws.max_row} cols={ws.max_column} ===')
    for row in ws.iter_rows(values_only=True):
        vals=['' if v is None else str(v).replace('\n',' / ') for v in row]
        if any(vals): print(' | '.join(vals))
